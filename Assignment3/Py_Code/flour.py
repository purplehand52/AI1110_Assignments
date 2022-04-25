#Import openyxl module
import openpyxl

#Open workbook
workbook = openpyxl.load_workbook("flour.xlsx")
worksheet = workbook.active

#Iterate loop
sample = []
for i in range(0, worksheet.max_row):
    for j in worksheet.iter_cols(1, worksheet.max_column):
        sample.append(j[i].value)

#n(Sample_Space)
n_s = len(sample)

#Random variable
n_x = 0
for i in sample:
    if i > 5:
        n_x = n_x + 1
    else:
        pass

#Probability
prob = n_x/n_s
print("Probability is", prob)
